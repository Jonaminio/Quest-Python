import datetime
from datetime import date
from openpyxl import load_workbook

nome = str(input('Digite seu nome: '))
tvacina = str(input('digite o tipo de vacina que deseja tomar, Pfizer ou Astrazeneca: '))
print('digite a data que esta tomando a primeira dose em ano/mês/dias: ')

year = int(input('Digite o ano: '))
month = int(input('Digite o mês: '))
day = int(input('Digite o dia: '))

date1 = datetime.date(year, month, day)

print(date1)

x = date1.strftime('%d/%m/%Y')
print(x)

if(tvacina == "pfizer"):
    futuro = date.fromordinal(date1.toordinal()+21)
else:
    futuro = date.fromordinal(date1.toordinal() + 28)

y = futuro.strftime('%d/%m/%Y')

print(f'Prezado(a) Sr(a): {nome},você tomou a vacina PFIZER, no dia {x} Sua segunda dose deverá ocorrer dia {y} ,Não se esqueça!')

############## salvando na planilha
caminho = 'Aprendendoalerplanilhas.xlsx'
arquivo_excel = load_workbook(caminho, data_only=True)
planilha = arquivo_excel.active

#botando conteúdo cabeçalho
planilha.cell(row=1, column=1, value='Nome do Paciente')
planilha.cell(row=1, column=2, value='Tipo de Vacina')
planilha.cell(row=1, column=3, value= 'Dia da primeira dose')
planilha.cell(row=1, column=4, value='Dia que deverá tomar a segunda dose')
#botando conteúdo
planilha.cell(row=2, column=1, value= nome)
planilha.cell(row=2, column=2, value= tvacina)
planilha.cell(row=2, column=3, value= x)
planilha.cell(row=2, column=4, value= y)

arquivo_excel.save("Aprendendoalerplanilhas.xlsx")

