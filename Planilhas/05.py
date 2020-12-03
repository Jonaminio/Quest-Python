from openpyxl import load_workbook

# https://letscode.com.br/blog/aprenda-a-integrar-python-e-excel
# https://openpyxl.readthedocs.io/en/stable/tutorial.html

# caminho = 'AULAS-EAD-REGISTRO-DE-ATIVIDADE.xlsx'
# lendo a planilha
caminho = 'Aprendendoalerplanilhas.xlsx'
# abrindo
arquivo_excel = load_workbook(caminho, data_only=True)
# setando na aba da planilha
planilha = arquivo_excel.active
# apontando na linha e coluna
sim = 0
nao = 0
#a coluna para ficar linear nos cabeçalhos
coluna =5

for y in range(1, 11): #Para de aluno 1 até aluno 10
    aluno = ["Aluno"]
    alunonum = [y] #aluno recebe 1,2,3.....10
    #str(alunonum)
    aluno += alunonum #concatenação das listas
    uniao = " ".join([str(_) for _ in aluno]) #transformando srt + int em string e tirando os [] etc...
    # sem isso ele ficaria assim: ['Aluno', 1] e não daria certo nas condicionais
    str(uniao) #convertendo pra string e ficará normal =======> Aluno 1
    print(uniao) #print pra conferir
    for x in range(1, 40): #aqui começa a contar das colunas B e C ate a linha 40 verificando
        conteudo = planilha.cell(row=x, column=2) #coluna B
        ponteirocont = planilha.cell(row=x, column=3)#coluna C
        if conteudo.value == uniao and ponteirocont.value == "SIM":
            sim += 1

        elif conteudo.value == uniao and ponteirocont.value == "NÃO":
            nao += 1
    #saindo do laço 2 > aqui add as informações nas colunas
    planilha.cell(row=7, column=coluna, value=uniao)
    planilha.cell(row=8, column=coluna, value=sim)
    planilha.cell(row=9, column=coluna, value=nao)
    coluna += 1 #coluna ++ para a mesma ir pro lado
    print(sim)
    print(nao)
    sim = 0 #zerando a contagem pra não acumular os sims e nãos, ficando apenas do seu respectivo
    nao = 0

#print(sim)
#print(nao)
# botando conteúdo
planilha.cell(row=8, column=4, value='SIM:')

planilha.cell(row=9, column=4, value='NAO:')

# salvando na planilha
arquivo_excel.save("Aprendendoalerplanilhas.xlsx")
