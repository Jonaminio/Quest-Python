from openpyxl import load_workbook

#https://letscode.com.br/blog/aprenda-a-integrar-python-e-excel
#https://openpyxl.readthedocs.io/en/stable/tutorial.html

#caminho = 'AULAS-EAD-REGISTRO-DE-ATIVIDADE.xlsx'
#lendo a planilha
caminho = 'Aprendendoalerplanilhas.xlsx'
#abrindo
arquivo_excel = load_workbook(caminho, data_only=True)
#setando na aba da planilha
planilha = arquivo_excel.active
#apontando na linha e coluna
conteudo = planilha["B5"]
#botando conteúdo
planilha.cell(row=7, column=4, value='Teste euris')
#salvando na planilha
arquivo_excel.save("Aprendendoalerplanilhas.xlsx")
print(conteudo.value)