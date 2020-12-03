from openpyxl import load_workbook

#https://letscode.com.br/blog/aprenda-a-integrar-python-e-excel
#https://openpyxl.readthedocs.io/en/stable/tutorial.html

#caminho = 'AULAS-EAD-REGISTRO-DE-ATIVIDADE.xlsx'
#lendo a planilha
caminho = 'Aprendendoalerplanilhas.xlsx'
#abrindo
arquivo_excel = load_workbook(caminho, data_only=True)
#setando na aba da planilha
planilha = arquivo_excel.active
#apontando na linha e coluna
sim =0
nao=0

colunas = 6

colunan = 6
xx = 6
xy = 6

planilha.cell(row=2, column=5, value='SIM:')
planilha.cell(row=3, column=5, value='NAO:')

for x in range(1,40):

    aluno = ["Aluno"]
    alunonum = [x]
    str(alunonum)
    aluno += alunonum
    uniao = " ".join([str(_) for _ in aluno])
    str(uniao)
    print(uniao)

    conteudo = planilha.cell(row=x, column=2)
    ponteirocont = planilha.cell(row=x, column=3)
    if conteudo.value == uniao and ponteirocont.value == "SIM":
        sim +=1
        planilha.cell(row=1, column=xx, value= uniao)
        planilha.cell(row=2, column=colunas, value=sim)
        colunas += 1
        xx += 1
    elif conteudo.value == uniao and ponteirocont.value == "NÃO":
        nao += 1
        planilha.cell(row=1, column=xy, value= uniao)
        planilha.cell(row=3, column=colunan, value= nao)
        colunan += 1
        xy += 1
    print(sim)
    print(nao)

#print(sim)
#print(nao)
#botando conteúdo
#planilha.cell(row=1, column=4, value="Aluno 5")
#planilha.cell(row=8, column=4, value='SIM:')
#planilha.cell(row=8, column=5, value= sim)
#planilha.cell(row=9, column=4, value='NAO:')
#planilha.cell(row=9, column=5, value= nao)
#salvando na planilha
arquivo_excel.save("Aprendendoalerplanilhas.xlsx")
print(conteudo.value)