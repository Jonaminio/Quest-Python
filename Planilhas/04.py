from openpyxl import load_workbook

# https://letscode.com.br/blog/aprenda-a-integrar-python-e-excel
# https://openpyxl.readthedocs.io/en/stable/tutorial.html

# caminho = 'AULAS-EAD-REGISTRO-DE-ATIVIDADE.xlsx'
# lendo a planilha
caminho = 'Aprendendoalerplanilhas.xlsx'
# abrindo
arquivo_excel = load_workbook(caminho, data_only=True)
# setando na aba da planilha
planilha = arquivo_excel.active
# apontando na linha e coluna
sim = 0
nao = 0
for y in range(1, 11):
    aluno = ["Aluno"]
    alunonum = [y]
    str(alunonum)
    aluno += alunonum
    uniao = " ".join([str(_) for _ in aluno])
    str(uniao)
    print(uniao)
    for x in range(1, 40):
        conteudo = planilha.cell(row=x, column=2)
        ponteirocont = planilha.cell(row=x, column=3)
        if conteudo.value == uniao and ponteirocont.value == "SIM":
            sim += 1
        elif conteudo.value == uniao and ponteirocont.value == "NÃO":
            nao += 1
    print(sim)
    print(nao)
    sim = 0
    nao = 0

#print(sim)
#print(nao)
# botando conteúdo
planilha.cell(row=7, column=4, value= uniao)
planilha.cell(row=8, column=4, value='SIM:')
planilha.cell(row=8, column=5, value=sim)
planilha.cell(row=9, column=4, value='NAO:')
planilha.cell(row=9, column=5, value=nao)
# salvando na planilha
arquivo_excel.save("Aprendendoalerplanilhas.xlsx")
